"""Class InfoWritter, python 3.8.1, tested on Win7 x64."""

class InfoWritter:
    """The class represents a mechanism that rewrites data from the KPP_WFRP program (info.info file) into a prepared worksheet in xlsx format."""
    #********************** constant ***********************
    __ATTR_ADDRESS = {"IMIE":(["B3"],1),
                    "RASA":(["R3"],1),
                    "PLEC":(["W3"],1),
                    "KLASA":(["AA3"],1),
                    "CHAR":(["AF3"],1),
                    "WIEK":(["B6"],1),
                    "WZROST":(["E6"],1),
                    "WAGA":(["I6"],1),
                    "WLOSY":(["M6"],1),
                    "OCZY":(["Q6"],1),
                    "OPIS":(["U6"],4),
                    "OB_PROF":(["B9"],1),
                    "WYJ_PROF":(["I9"],1),
                    "CHAR_POCZ":(["H12","J12","L12","N12","P12","R12","T12","V12","X12","Z12","AB12","AD12","AF12","AH12"],1),
                    "SCH_ROZW":(["H13","J13","L13","N13","P13","R13","T13","V13","X13","Z13","AB13","AD13","AF13","AH13"],1),
                    "CHAR_AKTUAL":(["H14","J14","L14","N14","P14","R14","T14","V14","X14","Z14","AB14","AD14","AF14","AH14"],1),
                    "BR_RECZNA":(["B17","J17","L17","N17","P17"],11),
                    "BR_STRZEL":(["B30","H30","J30","L30","N30","P30"],11),
                    "ZBROJA":(["B43","J43","P43"],11),
                    "UMIEJET":(["S17","AB17"],19),
                    "P_PANC":(["Y38","Y44","Y48","Y52","S52","S45","U41"],1),
                    "EQ":(["B3"],28),
                    "MAJATEK":(["B33"],3),
                    "TEMPO":(["X4","AA4","AD4","X6","AA6","AD6","X8","AA8","AD8"],1),
                    "JEZYKI":(["S13"],5),
                    "PSYCHA":(["X13"],5),
                    "OBLED":(["AC13"],5),
                    "M_URODZ":(["X19"],3),
                    "ZAW_RODZIC":(["X23"],3),
                    "RODZINA":(["S28"],5),
                    "POZ_SOC":(["U33"],1),
                    "RELIGIA":(["Y34"],2),
                    "P_P":(["AL38"],1),
                    "P_MAG":(["AL41"],1),
                    "P_MOCY":(["AL44"],1),
                    "TOWARZYSZE":(["B39","I39","K39","M39","O39","Q39","S39","U39","W39","Y39","AA39","AC39","AE39","AG39","AI39"],6),
                    "CZARY":(["B3","Q3","T3","W3","Z3","AD3","AO3"],25)}

    #separators contained in the info file
    __ATTR_SEP = ":"
    __DATA_SEP = ";"
    __EMPTY_SEP = "_"

    #control flags
    __DEBUG = False

    @staticmethod
    def change_flag():
        InfoWritter.__DEBUG = True

    def __init__(self, infoFileName, workbookObj):
        self.infoFile = infoFileName
        self.wb = workbookObj
        self.__characterName = None

    def __incDestAddrRow(self, dest_addr):
        """Returns the cell with the cell row value increased by 1"""
        #split destination addr on alpha and numbers
        r = ""
        c = ""
        for addr in dest_addr:
            for char in addr:
                if char.isalpha(): c += char
                else: r += char
        r = int(r)
        r += 1
        return c + str(r)

    def __conv2proper(self, value):
        """Returns value converted to proper type: str or int"""
        if value.isdigit(): return int(value)
        else: return value

    def __write2cell(self, dest_attr_address, data_, sheet):
        """Write data to specific cell in selected sheet in xlsx file. Inp ex: write2cell(what_addr, list_of_data_from_info, worksheet[0])"""
        #unpack dest address and how much rows is there
        #ex.: [B3], 1
        dest_addresses, _ = dest_attr_address
        data_ = list(map(str, data_))

        ncols = len(dest_addresses)
        ndata = len(data_)

        #if the number of columns corresponds to the number of data
        if ncols == ndata:
            for i, item in enumerate(data_):
                item = item.strip()
                if item == "0": item = ""
                sheet[dest_addresses[i]] = self.__conv2proper(item)
                if InfoWritter.__DEBUG:
                    print(f"\t{dest_addresses[i]} --> {item}")

        #if the number of columns is less than the number of data
        if ncols < ndata:
            i = 0
            for item in data_:
                item = item.strip()
                if i >= ncols:
                    dest_addresses = list(map(self.__incDestAddrRow, dest_addresses))
                    i = 0
                if item == "0": item = ""
                sheet[dest_addresses[i]] = self.__conv2proper(item)
                if InfoWritter.__DEBUG:
                    print(f"\t{dest_addresses[i]} --> {item}")
                i += 1

    def getCharName(self):
        return self.__characterName     

    def run(self):
        """Opens the info.info file and rewrites its formatted content to the specific cells in the xlsx file."""
        ws = self.wb["FRONT"]
        for nr, line in enumerate(open(self.infoFile, "r", encoding='utf-8')):
            if nr == 21: ws = self.wb["BACK"]
            if nr == 36: ws = self.wb["CZARY"]
            name, data = line.split(InfoWritter.__ATTR_SEP)
            if name == "IMIE": self.__characterName = data
            data = data.split(InfoWritter.__DATA_SEP)
            if InfoWritter.__EMPTY_SEP in data[0]: continue
            if InfoWritter.__DEBUG:
                print(f"\n{name}:")
            self.__write2cell(InfoWritter.__ATTR_ADDRESS[name], data, ws)


###############################################
#************* using the class ****************
if __name__ == "__main__":
    import sys
    from openpyxl import load_workbook

    def message(type):
        if type == 1:
            print("Something went wrong, maybe you entered the parameters incorrectly. Type in the console: \"wfrp2xlsx -help\" and press ENTER")
        elif type == 2:
            print("""The program transfers data from the input file to the output file.
    The input file comes from the KPP_WFRPed1_PL.exe program (v1.0.0.3) and is named: info.info
    The output file is a WFRP ed1 character card. Created in a spreadsheet. Must have the * .xlsx extension.

    Example of correct use of the program:
    ex.1: wfrp2xlsx -help   ---> shows the help content.
    ex.2: wfrp2xlsx -in info.info -out wfrp.xlsx    ---> correct call.
    ex.3: wfrp2xlsx -d -in info.info -out wfrp.xlsx    ---> correct call with debug mode.""")

    #possible params: "-help", "-in", "-out"
    #wfrp2xlsx.py -in info.info -out wfrp.xlsx
    if len(sys.argv) == 6:
        flag = False
        if len(sys.argv[5]) > 5 and sys.argv[5][-5:] == ".xlsx":
            flag = True
            if sys.argv[1] == "-d" and sys.argv[2] == "-in" and sys.argv[3][-5:] == ".info" and sys.argv[4] == "-out" and flag == True:
                #>>>> class work <<<<
                print("please wait...")
                workBook = load_workbook(sys.argv[5].strip())
                writter = InfoWritter(sys.argv[3].strip(), workBook)
                InfoWritter.change_flag() #<---- enabling debug mode
                writter.run()
                xlsxfilename = writter.getCharName().strip().replace(" ", "_") + ".xlsx"
                workBook.save(xlsxfilename)
                print("\n**** done ****")
            else:
                message(1)
                sys.exit(1)
    elif len(sys.argv) == 5:
        flag = False
        if len(sys.argv[4]) > 5 and sys.argv[4][-5:] == ".xlsx":
            flag = True
            if sys.argv[1] == "-in" and sys.argv[2][-5:] == ".info" and sys.argv[3] == "-out" and flag == True:
                #>>>> class work <<<<
                print("please wait...")
                workBook = load_workbook(sys.argv[4].strip())
                writter = InfoWritter(sys.argv[2].strip(), workBook)
                writter.run()
                xlsxfilename = writter.getCharName().strip().replace(" ", "_") + ".xlsx"
                workBook.save(xlsxfilename)
                print("\n**** done ****")
            else:
                message(1)
                sys.exit(1)
    elif len(sys.argv) == 2:
        if sys.argv[1] == "-help":
            message(2)
    else:
        message(1)